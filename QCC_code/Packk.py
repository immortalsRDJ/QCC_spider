import requests
from bs4 import BeautifulSoup
from lxml import etree
from urllib.parse import quote
import csv
import requests
import time
import random
from openpyxl import load_workbook
from JSreverse import get_header
companylist = []
list_all = []
error_num = 0
nul_num = 0
sharenum = [str(i) for i in range(1,31)]

headers = {'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36'
}
cookies = {
        "QCCSESSID": "be59338ca4e328b6532564bd88",
        "qcc_did": "6ff485cc-79b0-44df-b439-338d132d6e24"
    }



#Step1找公司
with open('deduplicate.csv') as f:
    f_csv = csv.reader(f)
    for row in f_csv:
        if row !=[] and row[0]!="None":
            companylist.append(str(row[0]))



for com in companylist[1:]:
    l = []
    company = quote(com)
    print(com)

    url = f'https://www.qcc.com/web/search?key={company}'

    headers = get_header(url)

    res = requests.get(url=url, headers=headers,cookies=cookies,timeout=(2, 2)).content

    tree = etree.HTML(res)
    url = (tree.xpath('//span[@class="copy-title"]/a/@href') + [''])[0]

    time.sleep(random.randint(15,20))
    l_item = []
    if url == '':

        l_item.append(com)
        wb = load_workbook(filename='/Users/clairemeng/Desktop/bacis_info_10k.20240605145823533.xlsx')
        ws = wb.active

        ws.append(l_item)
        print(str(com)+'搜寻无果，自动跳过')

        wb.save('/Users/clairemeng/desktop/bacis_info_10k.20240605145823533.xlsx')
        nul_num += 1
        if nul_num >= 20:
            print('Some error occurred, plz check.')
            break

        time.sleep(random.randint(3,5))

    else:

        print(url)

        headers = get_header(url)


        response = requests.get(url=url, cookies=cookies,headers=headers)

        response.encoding = 'utf-8'
        html = response.text

        soup = BeautifulSoup(html, 'html.parser')
        # print(soup)

        td_tags = soup.find_all('td')
        for td_tag in td_tags:
            data = td_tag.get_text().strip()
            print(data)
            l.append(data)






        print(l)
        try:
            ind = l.index("经营范围")
        except:
            ind = l.index("宗旨和业务范围")

        qccl = l[:ind+2]



        table_key = ['统一社会信用代码','企业名称', '法定代表人' ,'登记状态', '成立日期', '注册资本', '实缴资本', '组织机构代码', '工商注册号', '纳税人识别号', '企业类型', '营业期限' ,'纳税人资质', '人员规模','参保人数', '核准日期', '所属地区', '登记机关' ,'进出口企业代码', '国标行业', '英文名', '注册地址', '通信地址', '经营范围']
        d = {}
        d['运营单位'] = str(com)
        for key in table_key:
            if key in qccl:
                indd = qccl.index(key)

                d[key] = qccl[indd+1]
            else:
                d[key] = ""

        print(d)
        l_item = []
        for key in d:
            l_item.append(d[key])
        wb = load_workbook(filename='/Users/clairemeng/Desktop/bacis_info_10k.20240605145823533.xlsx')
        ws = wb.active

        ws.append(l_item)
        print('add completed')

        wb.save('/Users/clairemeng/desktop/bacis_info_10k.20240605145823533.xlsx')


        if "经营范围" not in l:
            d = {}
            d['公司名称'] = com
            l_item = []
            for key in d:
                l_item.append(d[key])
            wb = load_workbook(filename='/Users/clairemeng/Desktop/share_info_10k.xlsx')
            ws = wb.active

            ws.append(l_item)

            wb.save('/Users/clairemeng/desktop/share_info_10k.xlsx')
            continue

        ind = l.index("经营范围")
        qccl = l[ind + 2:]




        #print(qccl)

        keyname = ["公司名称", "序号", "股东名称", "持股比例", "认缴出资额(万元)", "认缴出资日期", "首次持股日期",
                   "关联产品/机构"]

        k = 0

        while qccl[7 * k] == sharenum[k]:

            d = {}
            c = 0
            d['公司名称'] = com
            for aaa in keyname[1:]:
                d[aaa] = qccl[7 * k + c]
                c += 1
            l_item = []
            for key in d:
                l_item.append(d[key])
            wb = load_workbook(filename='/Users/clairemeng/Desktop/share_info_10k.xlsx')
            ws = wb.active

            ws.append(l_item)

            wb.save('/Users/clairemeng/desktop/share_info_10k.xlsx')
            k += 1

        time.sleep(random.randint(20,25))